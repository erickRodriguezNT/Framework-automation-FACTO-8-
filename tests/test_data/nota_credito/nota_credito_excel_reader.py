"""
nota_credito_excel_reader.py — Lector de datos de prueba de Nota de Crédito desde Excel.

Responsabilidades:
- Cargar el archivo nota_credito_casos.xlsx (hoja "NotaCredito")
- Filtrar solo las filas marcadas con ejecutar = SI
- Normalizar los valores (NaN → "", SI/NO → bool, AUTO → fecha actual)
- Devolver una lista de dicts planos, uno por caso ejecutable

Uso:
    from tests.test_data.nota_credito.nota_credito_excel_reader import get_executable_nota_credito_cases
    casos = get_executable_nota_credito_cases()
    for caso in casos:
        print(caso["caso_id"], caso["rfc_receptor"])
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from app.utils.logger import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).parents[3]  # automation_framework/
_DEFAULT_EXCEL_PATH = (
    _PROJECT_ROOT / "tests" / "test_data" / "nota_credito" / "nota_credito_casos.xlsx"
)
_SHEET_NAME = "NotaCredito"

# ---------------------------------------------------------------------------
# Columnas booleanas: "SI" → True, "NO" / vacío → False
# ---------------------------------------------------------------------------
_BOOL_COLUMNS = {
    "ejecutar",
    "descargar_pdf",
    "descargar_xml",
}

# ---------------------------------------------------------------------------
# Columnas que deben mantenerse como string (incluso si pandas infiere número)
# ---------------------------------------------------------------------------
_STRING_COLUMNS = {
    "caso_id",
    "codigo_postal",
    "no_identificacion",
    "clave_prod_serv",
    "clave_unidad",
    "tasa_o_cuota",
    "tipo_cambio",
    "cantidad",
    "valor_unitario",
    "descuento",
    "cargo_no_facturable_importe",
    "forma_pago",
    "metodo_pago",
    "exportacion",
    "moneda",
    "objeto_impuesto",
    "impuesto",
    "uuid_relacionado",
    "tipo_relacion",
    "uuids_seleccionados",
}


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------


def load_nota_credito_cases_from_excel(
    path: str | Path | None = None,
    sheet_name: str = _SHEET_NAME,
) -> list[dict[str, Any]]:
    """
    Carga todos los casos del Excel de Nota de Crédito (ejecutables o no).

    Args:
        path:       Ruta al archivo .xlsx. Si None, usa la ruta por defecto.
        sheet_name: Nombre de la hoja. Default: "NotaCredito".

    Returns:
        Lista de dicts, uno por fila del Excel. Valores normalizados.

    Raises:
        FileNotFoundError: Si el archivo no existe.
        ValueError:        Si la hoja no existe o hay error al leer.
    """
    try:
        import pandas as pd  # lazy import
    except ImportError as exc:
        raise ImportError(
            "pandas no está instalado. Ejecutar: pip install pandas openpyxl"
        ) from exc

    excel_path = Path(path) if path else _DEFAULT_EXCEL_PATH

    if not excel_path.exists():
        raise FileNotFoundError(
            f"Archivo de casos Excel no encontrado: {excel_path}\n"
            f"Crear el archivo en: tests/test_data/nota_credito/nota_credito_casos.xlsx\n"
            f"Ejecutar: python tests/test_data/nota_credito/nota_credito_excel_reader.py"
        )

    logger.info(f"[NC EXCEL] Cargando casos desde: {excel_path}")

    try:
        df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name,
            dtype=str,           # leer TODO como string; normalizaremos manualmente
            keep_default_na=False,  # NaN → "" en lugar de float NaN
        )
    except Exception as exc:
        raise ValueError(f"Error al leer el Excel '{excel_path}': {exc}") from exc

    if df.empty:
        logger.warning("[NC EXCEL] El Excel no tiene filas de datos.")
        return []

    # Normalizar nombres de columnas: strip espacios
    df.columns = [str(c).strip() for c in df.columns]

    logger.info(f"[NC EXCEL] {len(df)} filas cargadas. Columnas: {list(df.columns)}")

    casos: list[dict[str, Any]] = []
    for idx, row in df.iterrows():
        caso = _normalize_row(row.to_dict(), idx)
        casos.append(caso)

    return casos


def get_executable_nota_credito_cases(
    path: str | Path | None = None,
    sheet_name: str = _SHEET_NAME,
) -> list[dict[str, Any]]:
    """
    Carga los casos del Excel y retorna solo los marcados como ejecutables.

    Regla: solo se ejecutan filas donde la columna 'ejecutar' = "SI" (case-insensitive).

    Args:
        path:       Ruta al archivo .xlsx. Si None, usa la ruta por defecto.
        sheet_name: Nombre de la hoja. Default: "NotaCredito".

    Returns:
        Lista de dicts de casos ejecutables. Puede ser vacía.
    """
    todos = load_nota_credito_cases_from_excel(path=path, sheet_name=sheet_name)
    ejecutables = [c for c in todos if c.get("ejecutar") is True]
    logger.info(
        f"[NC EXCEL] {len(ejecutables)} / {len(todos)} casos marcados como ejecutables."
    )
    return ejecutables


# ---------------------------------------------------------------------------
# Helpers de normalización
# ---------------------------------------------------------------------------


def _normalize_row(raw: dict[str, Any], idx: int) -> dict[str, Any]:
    """
    Normaliza una fila del Excel a un dict listo para el flujo de Nota de Crédito.

    Transformaciones aplicadas:
    - Todas las claves: strip de espacios
    - Todos los valores: strip de espacios y convertir NaN/None a ""
    - Columnas bool (_BOOL_COLUMNS): "SI" → True, todo lo demás → False
    - fecha_emision = "AUTO" → fecha/hora actual en formato datetime-local
    - referencia = "AUTO" → referencia con timestamp
    - caso_id vacío → generar ID por índice: "NC_{idx+1}"
    - uuid_relacionado vacío → "" (opcional, no requerido en esta etapa)
    - tipo_relacion vacío → "" (opcional, no requerido en esta etapa)

    Args:
        raw: Diccionario crudo de la fila.
        idx: Índice de la fila (0-based).

    Returns:
        Diccionario normalizado listo para usar en NotaCreditoFlow.
    """
    normalized: dict[str, Any] = {}

    for key, value in raw.items():
        key = str(key).strip()

        # Normalizar valor: NaN, None, "nan", "NaN" → ""
        if value is None or (isinstance(value, float) and __import__("math").isnan(value)):
            value = ""
        else:
            value = str(value).strip()
            if value.lower() == "nan":
                value = ""

        # Columnas booleanas
        if key in _BOOL_COLUMNS:
            normalized[key] = value.upper() in ("SI", "SÍ", "YES", "TRUE", "1")
            continue

        # fecha_emision AUTO → fecha actual compatible con datetime-local
        if key == "fecha_emision":
            if value.upper() == "AUTO" or value == "":
                normalized[key] = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                logger.debug(f"[NC EXCEL] fecha_emision AUTO → {normalized[key]}")
            else:
                normalized[key] = value
            continue

        # referencia AUTO → generar referencia con timestamp
        if key == "referencia" and value.upper() == "AUTO":
            normalized[key] = f"AUTO-NC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            continue

        # no_identificacion AUTO → generar id con timestamp
        if key == "no_identificacion" and value.upper() == "AUTO":
            normalized[key] = f"AUTO-NC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            continue

        normalized[key] = value

    # caso_id vacío → ID por índice con prefijo NC
    if not normalized.get("caso_id"):
        normalized["caso_id"] = f"NC_{idx + 1}"
        logger.warning(
            f"[NC EXCEL] Fila {idx + 1} sin caso_id. Asignado automáticamente: {normalized['caso_id']}"
        )

    # Asegurar que campos CFDI relacionado existen como string vacío si ausentes
    normalized.setdefault("tipo_relacion", "")
    normalized.setdefault("uuids_seleccionados", "")
    normalized.setdefault("uuid_relacionado", "")

    return normalized


# ---------------------------------------------------------------------------
# Script de generación del Excel de ejemplo
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    """
    Genera el archivo nota_credito_casos.xlsx con todas las columnas requeridas
    y una fila de ejemplo lista para ejecutar.

    Ejecutar desde la raíz del framework:
        python tests/test_data/nota_credito/nota_credito_excel_reader.py
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("ERROR: openpyxl no está instalado. Ejecutar: pip install openpyxl")
        raise SystemExit(1)

    # Columnas requeridas según la especificación del proyecto
    COLUMNAS = [
        "ejecutar",
        "caso_id",
        "descripcion",
        "centro_consumo",
        "serie",
        "rfc_receptor",
        "razon_social",
        "codigo_postal",
        "regimen_fiscal_receptor",
        "uso_cfdi",
        "email",
        "calle",
        "num_exterior",
        "num_interior",
        "colonia",
        "municipio",
        "estado",
        "fecha_emision",
        "exportacion",
        "moneda",
        "tipo_cambio",
        "metodo_pago",
        "forma_pago",
        "condiciones_pago",
        "observaciones_pdf",
        "idioma",
        "integraciones",
        "referencia",
        "clave_unidad",
        "cantidad",
        "no_identificacion",
        "clave_prod_serv",
        "descripcion_concepto",
        "valor_unitario",
        "descuento",
        "objeto_impuesto",
        "impuesto",
        "ret_tras",
        "tipo_factor",
        "tasa_o_cuota",
        "cargo_no_facturable_nombre",
        "cargo_no_facturable_importe",
        "descargar_pdf",
        "descargar_xml",
        "tipo_relacion",
        "uuids_seleccionados",
        "uuid_relacionado",
        "resultado_esperado",
    ]

    # Fila de ejemplo: NC_001
    EJEMPLO = {
        "ejecutar": "SI",
        "caso_id": "NC_001",
        "descripcion": "Nota de crédito base replicando flujo de factura manual",
        "centro_consumo": "NID200929V26 - PRUEBAS - Pruebas",
        "serie": "",
        "rfc_receptor": "XAXX010101000",
        "razon_social": "PUBLICO EN GENERAL",
        "codigo_postal": "03104",
        "regimen_fiscal_receptor": "616 - Sin obligaciones fiscales",
        "uso_cfdi": "S01 - Sin efectos fiscales",
        "email": "alejandro.rodriguez@next-technologies.com.mx",
        "calle": "",
        "num_exterior": "",
        "num_interior": "",
        "colonia": "",
        "municipio": "",
        "estado": "",
        "fecha_emision": "AUTO",
        "exportacion": "01 - No aplica",
        "moneda": "MXN - Peso Mexicano",
        "tipo_cambio": "1",
        "metodo_pago": "PUE",
        "forma_pago": "01",
        "condiciones_pago": "Contado",
        "observaciones_pdf": "Nota de crédito generada por automatización",
        "idioma": "Español",
        "integraciones": "",
        "referencia": "AUTO",
        "clave_unidad": "E48",
        "cantidad": "1",
        "no_identificacion": "AUTO",
        "clave_prod_serv": "84111506",
        "descripcion_concepto": "AJUSTE POR NOTA DE CREDITO",
        "valor_unitario": "1000.00",
        "descuento": "0.00",
        "objeto_impuesto": "02",
        "impuesto": "002",
        "ret_tras": "Traslado",
        "tipo_factor": "Tasa",
        "tasa_o_cuota": "0.160000",
        "cargo_no_facturable_nombre": "",
        "cargo_no_facturable_importe": "",
        "descargar_pdf": "SI",
        "descargar_xml": "SI",
        "tipo_relacion": "01 - Nota de crédito de los documentos relacionados",
        "uuids_seleccionados": "",
        "uuid_relacionado": "",
        "resultado_esperado": "TIMBRADO_EXITOSO",
    }

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NotaCredito"

    # Estilos de encabezado
    header_fill = PatternFill(start_color="1760D7", end_color="1760D7", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escribir encabezados
    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 15)

    # Escribir fila de ejemplo
    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        ws.cell(row=2, column=col_idx, value=EJEMPLO.get(col_name, ""))

    # Altura de encabezado
    ws.row_dimensions[1].height = 40

    # Guardar archivo
    output_path = _DEFAULT_EXCEL_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"[OK] Archivo generado: {output_path}")
    print(f"     Hoja: NotaCredito")
    print(f"     Columnas: {len(COLUMNAS)}")
    print(f"     Filas de ejemplo: 1 (NC_001 con ejecutar=SI)")
